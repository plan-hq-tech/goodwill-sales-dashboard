
import io
import re
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="굿윌 매출 리포트", page_icon="📈", layout="wide")

MONTH_PATTERNS = [
    re.compile(r"(20\d{2})[-_]?([01]?\d)"),
    re.compile(r"(20\d{2})년\s*([01]?\d)월"),
]

DAILY_NUM_COLS = ["전표건수","객단가","공급가액","실매출액","현금","현금영수증","카드","포인트","현금카드외","제휴포인트","상품권결제"]
PRODUCT_NUM_COLS = ["판매수량","공급가액","실매출액","기본단가","이익금액","현금","현금영수증","카드","상품권결제","현금카드외"]

MAJOR_DONORS = ["CJ제일제당","편의점","모던하우스","오뚜기","신세계푸드"]
DONOR_PATTERNS = {
    "CJ제일제당": [r"CJ제일제당", r"\bCJ\b"],
    "편의점": [r"GS25", r"\bCU\b", r"세븐일레븐", r"편의점"],
    "모던하우스": [r"모던하우스", r"기증파트너 》 모던", r"\b모던\b"],
    "오뚜기": [r"오뚜기"],
    "신세계푸드": [r"신세계푸드"],
}
CATEGORY_ORDER = ["의류","잡화","생활","식품","건강/미용","문화","원가상품","기타"]

# style
TITLE_FILL = PatternFill("solid", fgColor="1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
HEADER_BLUE = PatternFill("solid", fgColor="DDEBF7")
HEADER_YELLOW = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=14)
BOLD_FONT = Font(bold=True)
RED_FONT = Font(color="C00000", bold=True)
BLUE_FONT = Font(color="1F4E78", bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

def clean_number(x):
    if pd.isna(x):
        return 0.0
    if isinstance(x, (int, float, np.number)):
        return float(x)
    s = str(x).strip().replace(",", "").replace("%", "")
    s = s.replace("합계:", "").replace("평균:", "").replace("카운트:", "")
    if s == "":
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

def pct(a, b):
    if b in [0, None] or pd.isna(b):
        return 0.0
    return float(a) / float(b)

def fmt_won(x): return f"{x:,.0f}원"
def fmt_pct(x): return f"{x*100:,.1f}%"

def extract_month_from_filename(filename: str) -> Optional[str]:
    name = Path(filename).stem
    for pattern in MONTH_PATTERNS:
        m = pattern.search(name)
        if m:
            y = int(m.group(1))
            mm = int(m.group(2))
            if 1 <= mm <= 12:
                return f"{y:04d}-{mm:02d}"
    return None

def month_label(month_str: str) -> str:
    y, m = month_str.split("-")
    return f"{y}년 {int(m)}월"

def auto_fit(ws, min_width=9, max_width=24):
    for col_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            length = max(length, len(val))
        ws.column_dimensions[col_letter].width = max(min(length + 2, max_width), min_width)

def style_title(ws, row, end_col, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = TITLE_FILL
    c.font = WHITE_FONT
    c.alignment = CENTER

def style_section(ws, row, end_col, title, fill=SECTION_FILL):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = fill
    c.font = BOLD_FONT
    c.alignment = LEFT

def apply_table_style(ws, start_row, end_row, start_col=1, end_col=None, header_fill=HEADER_FILL,
                      pct_cols=None, won_cols=None, int_cols=None):
    end_col = end_col or ws.max_column
    pct_cols = pct_cols or []
    won_cols = won_cols or []
    int_cols = int_cols or []
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            if r == start_row:
                cell.fill = header_fill
                cell.font = BOLD_FONT
                cell.alignment = CENTER
            else:
                cell.alignment = CENTER if c > 1 else LEFT
                if c in pct_cols:
                    cell.number_format = '0.0%'
                elif c in won_cols:
                    cell.number_format = '#,##0'
                elif c in int_cols:
                    cell.number_format = '#,##0'

def parse_daily_sales(uploaded_file):
    raw = pd.read_excel(uploaded_file, sheet_name=0)
    raw = raw.rename(columns=lambda x: str(x).strip())
    store_col = raw.columns[0]
    current_store = None
    rows = []
    for _, row in raw.iterrows():
        marker = row.get(store_col)
        if isinstance(marker, str) and "매장:" in marker:
            current_store = re.sub(r".*매장:\s*", "", marker).split("[")[0].strip().replace("밀알","")
            continue
        sale_date = row.get("영업일자")
        if pd.isna(sale_date):
            continue
        try:
            sale_date = pd.to_datetime(sale_date)
        except Exception:
            continue
        out = {"지점명": current_store if current_store else "미확인", "영업일자": sale_date, "기준월": sale_date.strftime("%Y-%m")}
        for col in DAILY_NUM_COLS:
            out[col] = clean_number(row.get(col))
        rows.append(out)
    df = pd.DataFrame(rows)
    if not df.empty:
        df["일"] = df["영업일자"].dt.day
        df["월"] = df["영업일자"].dt.month
        df["연도"] = df["영업일자"].dt.year
        df["객단가_계산"] = df.apply(lambda r: pct(r["실매출액"], r["전표건수"]) if r["전표건수"] else 0, axis=1)
    return df

def infer_donor(row):
    text = f"{row.get('상품명','')} | {row.get('상품분류','')}"
    for donor, patterns in DONOR_PATTERNS.items():
        for p in patterns:
            if re.search(p, text, re.IGNORECASE):
                return donor
    if "》" in str(row.get("상품분류","")):
        parts = [x.strip() for x in str(row.get("상품분류","")).split("》")]
        if len(parts) >= 2:
            second = parts[1]
            blacklist = {"여성의류","남성의류","아동의류","하의","상의","가방","신발","잡화","문화용품","도서","생활용품","주방용품","가전","건강/미용","기업","식품","매입","제빵"}
            if second not in blacklist and len(second) >= 2:
                return second
    return "기타"

def parse_product_sales(uploaded_file):
    raw = pd.read_excel(uploaded_file, sheet_name=0)
    raw = raw.rename(columns=lambda x: str(x).strip())
    month_str = extract_month_from_filename(uploaded_file.name)
    if not month_str:
        raise ValueError(f"상품별 파일명에서 월을 읽을 수 없습니다: {uploaded_file.name}")
    category_hint = None
    rows = []
    for _, row in raw.iterrows():
        left = row.get("매장")
        if isinstance(left, str) and "상품분류1:" in left:
            category_hint = re.sub(r".*상품분류1:\s*", "", left).split("[")[0].strip()
            continue
        store = row.get("Unnamed: 1")
        product = row.get("상품")
        if pd.isna(store) or pd.isna(product):
            continue
        category = str(row.get("상품분류")).strip() if not pd.isna(row.get("상품분류")) else (category_hint or "미분류")
        out = {
            "기준월": month_str,
            "지점명": str(store).strip().replace("밀알",""),
            "상품명": str(product).strip(),
            "상품분류": category,
        }
        for col in PRODUCT_NUM_COLS:
            out[col] = clean_number(row.get(col))
        rows.append(out)
    df = pd.DataFrame(rows)
    if not df.empty:
        df["대분류"] = df["상품분류"].astype(str).str.split("》").str[0].str.strip()
        df["기증처"] = df.apply(infer_donor, axis=1)
    return df

@st.cache_data(show_spinner=False)
def load_all_data(daily_files, product_files):
    daily = pd.concat([parse_daily_sales(f) for f in daily_files], ignore_index=True) if daily_files else pd.DataFrame()
    product = pd.concat([parse_product_sales(f) for f in product_files], ignore_index=True) if product_files else pd.DataFrame()
    return daily, product

def build_month_store_summary(daily):
    if daily.empty:
        return pd.DataFrame()
    out = daily.groupby(["기준월","지점명"], as_index=False).agg({
        "전표건수":"sum","공급가액":"sum","실매출액":"sum",
        "현금":"sum","현금영수증":"sum","카드":"sum","포인트":"sum","현금카드외":"sum","제휴포인트":"sum","상품권결제":"sum"
    })
    out["객단가"] = out.apply(lambda r: pct(r["실매출액"], r["전표건수"]) if r["전표건수"] else 0, axis=1)
    out = out.sort_values(["지점명","기준월"])
    out["전월실매출액"] = out.groupby("지점명")["실매출액"].shift(1)
    out["전월대비증감률"] = out.apply(lambda r: pct(r["실매출액"]-r["전월실매출액"], r["전월실매출액"]) if r["전월실매출액"] else 0, axis=1)
    return out

def normalize_category_group(x):
    x = str(x)
    if "의류" in x: return "의류"
    if "잡화" in x: return "잡화"
    if "생활용품" in x or "생활" in x: return "생활"
    if "식품" in x: return "식품"
    if "건강/미용" in x or "건강" in x or "미용" in x: return "건강/미용"
    if "문화" in x or "도서" in x: return "문화"
    if "원가상품" in x or "매입" in x: return "원가상품"
    return "기타"

def build_classification_report(product_month):
    if product_month.empty:
        return pd.DataFrame()
    df = product_month.copy()
    df["분류그룹"] = df["대분류"].apply(normalize_category_group)
    grp = df.groupby(["지점명","분류그룹"], as_index=False).agg({"판매수량":"sum","실매출액":"sum"})
    totals = grp.groupby("지점명", as_index=False)["실매출액"].sum().rename(columns={"실매출액":"지점합계"})
    grp = grp.merge(totals, on="지점명", how="left")
    grp["점유율"] = grp.apply(lambda r: pct(r["실매출액"], r["지점합계"]), axis=1)
    grp["정렬"] = grp["분류그룹"].apply(lambda x: CATEGORY_ORDER.index(x) if x in CATEGORY_ORDER else 999)
    return grp.sort_values(["지점명","정렬"]).drop(columns=["정렬","지점합계"])

def build_donor_report(product_month):
    if product_month.empty:
        return pd.DataFrame()
    grp = product_month.groupby(["지점명","기증처"], as_index=False).agg({"판매수량":"sum","실매출액":"sum"})
    grp["피스단가"] = grp.apply(lambda r: pct(r["실매출액"], r["판매수량"]) if r["판매수량"] else 0, axis=1)
    base = grp[grp["기증처"].isin(MAJOR_DONORS)].copy()
    extra = (
        grp[~grp["기증처"].isin(MAJOR_DONORS)]
        .groupby("기증처", as_index=False)["실매출액"].sum()
        .sort_values("실매출액", ascending=False)
        .head(5)
    )
    extra_names = extra["기증처"].tolist()
    extra_df = grp[grp["기증처"].isin(extra_names)].copy()
    out = pd.concat([base, extra_df], ignore_index=True)
    total_order = out.groupby("기증처", as_index=False)["실매출액"].sum().sort_values("실매출액", ascending=False)["기증처"].tolist()
    out["기증처"] = pd.Categorical(out["기증처"], categories=total_order, ordered=True)
    return out.sort_values(["기증처","지점명"])

def build_same_month_last_year(current_df):
    if current_df.empty:
        return pd.DataFrame()
    months = current_df["기준월"].dropna().unique().tolist()
    result = []
    for month in months:
        y, m = month.split("-")
        prev = f"{int(y)-1:04d}-{m}"
        curr = current_df[current_df["기준월"] == month].copy()
        prev_df = current_df[current_df["기준월"] == prev].copy()
        if curr.empty:
            continue
        curr = curr.rename(columns={"판매수량":"판매수량_당해","실매출액":"실매출액_당해","점유율":"점유율_당해"})
        prev_df = prev_df.rename(columns={"판매수량":"판매수량_전년","실매출액":"실매출액_전년","점유율":"점유율_전년"})
        merged = curr.merge(prev_df[["지점명","분류그룹","판매수량_전년","실매출액_전년","점유율_전년"]],
                            on=["지점명","분류그룹"], how="left")
        merged["기준월"] = month
        result.append(merged)
    return pd.concat(result, ignore_index=True) if result else pd.DataFrame()

def monthly_overview_table(month_store, month):
    sub = month_store[month_store["기준월"] == month].copy().sort_values("실매출액", ascending=False)
    sub["전년대비"] = sub["전월대비증감률"]
    sub["전년매출"] = sub["전월실매출액"].fillna(0)
    sub["영업일수"] = np.nan
    sub["일평균매출"] = np.nan
    cols = ["지점명","실매출액","전표건수","객단가","전년대비","전년매출","영업일수","일평균매출"]
    out = sub[cols].copy()
    out.columns = ["매장","총매출","영수건수","객단가","전년대비","전년매출","영업일수","일평균 매출"]
    return out

def category_yoy_table(product_df, month):
    current = build_classification_report(product_df[product_df["기준월"] == month].copy())
    if current.empty:
        return pd.DataFrame()
    current_total = current.groupby("분류그룹", as_index=False).agg({"판매수량":"sum","실매출액":"sum"})
    current_total["점유율"] = current_total["실매출액"] / current_total["실매출액"].sum() if current_total["실매출액"].sum() else 0

    y, m = month.split("-")
    prev_month = f"{int(y)-1:04d}-{m}"
    prev = build_classification_report(product_df[product_df["기준월"] == prev_month].copy())
    prev_total = prev.groupby("분류그룹", as_index=False).agg({"판매수량":"sum","실매출액":"sum"}) if not prev.empty else pd.DataFrame(columns=["분류그룹","판매수량","실매출액"])
    if not prev_total.empty:
        prev_total["점유율"] = prev_total["실매출액"] / prev_total["실매출액"].sum() if prev_total["실매출액"].sum() else 0

    merged = current_total.merge(prev_total, on="분류그룹", how="outer", suffixes=("_당해","_전년")).fillna(0)
    merged["판매수량_차이"] = merged["판매수량_당해"] - merged["판매수량_전년"]
    merged["실매출액_차이"] = merged["실매출액_당해"] - merged["실매출액_전년"]
    merged["점유율_차이"] = merged["점유율_당해"] - merged["점유율_전년"]
    merged = merged.rename(columns={"분류그룹":"구분"})
    merged["정렬"] = merged["구분"].apply(lambda x: CATEGORY_ORDER.index(x) if x in CATEGORY_ORDER else 999)
    return merged.sort_values("정렬").drop(columns=["정렬"])

def receipt_comparison_table(month_store, month):
    y, m = month.split("-")
    prev = f"{int(y)-1:04d}-{m}"
    curr = month_store[month_store["기준월"] == month].copy()
    prev_df = month_store[month_store["기준월"] == prev].copy()
    merged = curr.merge(prev_df[["지점명","실매출액","전표건수","객단가"]], on="지점명", how="left", suffixes=("_당해","_전년"))
    merged["연도"] = int(y)
    merged["영수건수 증가수"] = merged["전표건수_당해"] - merged["전표건수_전년"].fillna(0)
    out = merged[["지점명","실매출액_당해","전표건수_당해","객단가_당해","연도","영수건수 증가수"]].copy()
    out.columns = ["매장","총매출","영수건수","객단가","연도","영수건수 증가수"]
    return out.sort_values("총매출", ascending=False)

def payment_mix_table(daily_month):
    cols = ["현금","현금영수증","카드","포인트","현금카드외","제휴포인트","상품권결제"]
    s = daily_month[cols].sum().reset_index()
    s.columns = ["결제수단","금액"]
    s = s[s["금액"] > 0].sort_values("금액", ascending=False)
    total = s["금액"].sum()
    s["점유율"] = s["금액"] / total if total else 0
    return s

def top_bottom_stores_table(month_store, month, n=5):
    sub = month_store[month_store["기준월"] == month].copy()
    top = sub.nlargest(n, "실매출액")[["지점명","실매출액","전표건수","객단가"]]
    bottom = sub.nsmallest(n, "실매출액")[["지점명","실매출액","전표건수","객단가"]]
    return top, bottom

def make_report_book(product_df, daily_df):
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        month_store = build_month_store_summary(daily_df)
        months = sorted(month_store["기준월"].unique().tolist())

        # reference sheet with image
        pd.DataFrame({"안내":["참고 양식 이미지"]}).to_excel(writer, sheet_name="참고양식", index=False)
        ws_ref = writer.book["참고양식"]
        style_title(ws_ref, 1, 6, "사용자 제공 참고 이미지")
        img_path = Path(__file__).parent / "sample_layout.png"
        if img_path.exists():
            img = XLImage(str(img_path))
            img.width = 900
            img.height = 1400
            ws_ref.add_image(img, "A3")

        if not months:
            out.seek(0)
            return out.getvalue()

        latest = months[-1]
        overview = monthly_overview_table(month_store, latest)
        category_yoy = category_yoy_table(product_df, latest)
        receipt_cmp = receipt_comparison_table(month_store, latest)
        payment_mix = payment_mix_table(daily_df[daily_df["기준월"] == latest].copy())
        top5, bottom5 = top_bottom_stores_table(month_store, latest)

        # 요약대시보드
        overview.to_excel(writer, sheet_name="요약대시보드", index=False, startrow=2)
        ws = writer.book["요약대시보드"]
        style_title(ws, 1, max(8, overview.shape[1]), f"{month_label(latest)} 요약대시보드")
        apply_table_style(ws, 3, 3 + len(overview), won_cols=[2,4,6,8], int_cols=[3])
        auto_fit(ws)
        ws.freeze_panes = "A4"

        # category yoy
        if not category_yoy.empty:
            category_yoy.to_excel(writer, sheet_name="분류전년비교", index=False, startrow=2)
            ws = writer.book["분류전년비교"]
            style_title(ws, 1, category_yoy.shape[1], f"{month_label(latest)} 분류별 전년 비교")
            apply_table_style(ws, 3, 3 + len(category_yoy), won_cols=[3,6,9], int_cols=[2,5,8], pct_cols=[4,7,10])
            auto_fit(ws)
            ws.freeze_panes = "A4"

        receipt_cmp.to_excel(writer, sheet_name="영수건수비교", index=False, startrow=2)
        ws = writer.book["영수건수비교"]
        style_title(ws, 1, receipt_cmp.shape[1], f"{month_label(latest)} 영수건수 비교")
        apply_table_style(ws, 3, 3 + len(receipt_cmp), won_cols=[2,4], int_cols=[3,6])
        auto_fit(ws)
        ws.freeze_panes = "A4"

        payment_mix.to_excel(writer, sheet_name="결제수단분석", index=False, startrow=2)
        ws = writer.book["결제수단분석"]
        style_title(ws, 1, payment_mix.shape[1], f"{month_label(latest)} 결제수단 분석")
        apply_table_style(ws, 3, 3 + len(payment_mix), won_cols=[2], pct_cols=[3])
        auto_fit(ws)

        # top bottom combined
        row = 1
        top5.to_excel(writer, sheet_name="상하위점포", index=False, startrow=row+2)
        bottom5.to_excel(writer, sheet_name="상하위점포", index=False, startrow=row+2, startcol=7)
        ws = writer.book["상하위점포"]
        style_title(ws, 1, 11, f"{month_label(latest)} 상위/하위 점포")
        style_section(ws, 3, 4, "상위 5개 점포", HEADER_BLUE)
        style_section(ws, 3, 10, "하위 5개 점포", HEADER_YELLOW)
        apply_table_style(ws, 4, 4 + len(top5), start_col=1, end_col=4, won_cols=[2,4], int_cols=[3], header_fill=HEADER_BLUE)
        apply_table_style(ws, 4, 4 + len(bottom5), start_col=8, end_col=11, won_cols=[9,11], int_cols=[10], header_fill=HEADER_YELLOW)
        auto_fit(ws)

    out.seek(0)
    return out.getvalue()

# -----------------------------
# UI
# -----------------------------
st.title("굿윌 매출 리포트")
st.caption("대시보드 분석 확대 + 참고 이미지 포함 보고서형 엑셀 다운로드")

with st.sidebar:
    st.header("파일 업로드")
    daily_files = st.file_uploader("매출현황 파일", type=["xlsx","xls"], accept_multiple_files=True)
    product_files = st.file_uploader("상품별 매출현황 파일", type=["xlsx","xls"], accept_multiple_files=True, help="파일명에 2026-04 또는 202604 형식의 월 표기 필요")

if not daily_files:
    st.info("먼저 매출현황 파일을 업로드해 주세요.")
    st.stop()

daily_df, product_df = load_all_data(daily_files, product_files)
if daily_df.empty:
    st.warning("매출현황 파일을 읽을 수 없습니다.")
    st.stop()

month_store = build_month_store_summary(daily_df)
months = sorted(month_store["기준월"].unique().tolist())
selected_month = st.sidebar.selectbox("기준월", months, index=len(months)-1)
stores = sorted(month_store["지점명"].unique().tolist())
selected_stores = st.sidebar.multiselect("지점 선택", stores, default=stores)

fm = month_store[(month_store["기준월"] == selected_month) & (month_store["지점명"].isin(selected_stores))].copy()
dm = daily_df[(daily_df["기준월"] == selected_month) & (daily_df["지점명"].isin(selected_stores))].copy()
pm = product_df[(product_df["기준월"] == selected_month) & (product_df["지점명"].isin(selected_stores))].copy() if not product_df.empty else pd.DataFrame()

total_sales = fm["실매출액"].sum()
total_cnt = fm["전표건수"].sum()
avg_ticket = pct(total_sales, total_cnt)
prev_sales = fm["전월실매출액"].fillna(0).sum() if "전월실매출액" in fm.columns else 0
mom = pct(total_sales - prev_sales, prev_sales) if prev_sales else 0

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("실매출액", fmt_won(total_sales))
c2.metric("전표건수", f"{total_cnt:,.0f}건")
c3.metric("객단가", fmt_won(avg_ticket))
c4.metric("전월대비", fmt_pct(mom))
c5.metric("점포수", f"{fm['지점명'].nunique():,}개")

tab1, tab2, tab3, tab4 = st.tabs(["요약", "지점 분석", "상품/기증처", "엑셀 다운로드"])

with tab1:
    left, right = st.columns([1.3,1])
    with left:
        st.subheader("월별 실매출 추이")
        trend = month_store[month_store["지점명"].isin(selected_stores)].groupby("기준월", as_index=False)["실매출액"].sum()
        st.line_chart(trend.set_index("기준월")["실매출액"])
        st.subheader("일별 실매출 추이")
        daily_trend = dm.groupby("영업일자", as_index=False)["실매출액"].sum().sort_values("영업일자")
        st.area_chart(daily_trend.set_index("영업일자")["실매출액"])
    with right:
        st.subheader("결제수단 비중")
        pay = payment_mix_table(dm)
        if not pay.empty:
            st.dataframe(pay, use_container_width=True, hide_index=True)
        st.subheader("상위 / 하위 점포")
        top5, bottom5 = top_bottom_stores_table(month_store, selected_month)
        a, b = st.columns(2)
        a.write("상위 5개")
        a.dataframe(top5, use_container_width=True, hide_index=True)
        b.write("하위 5개")
        b.dataframe(bottom5, use_container_width=True, hide_index=True)

with tab2:
    st.subheader("지점별 성과")
    view = fm[["지점명","실매출액","전표건수","객단가","전월대비증감률"]].sort_values("실매출액", ascending=False).copy()
    st.dataframe(view, use_container_width=True, hide_index=True)
    st.subheader("영수건수 기준 비교")
    st.dataframe(receipt_comparison_table(month_store, selected_month), use_container_width=True, hide_index=True)

with tab3:
    if pm.empty:
        st.info("상품별 파일을 업로드하면 분류별, 기증처별, 전년 비교가 표시됩니다.")
    else:
        class_df = build_classification_report(pm)
        donor_df = build_donor_report(pm)
        yoy = category_yoy_table(product_df, selected_month)

        a, b = st.columns(2)
        with a:
            st.subheader("분류별 매출 구성")
            st.bar_chart(class_df.groupby("분류그룹", as_index=False)["실매출액"].sum().set_index("분류그룹"))
            st.dataframe(class_df, use_container_width=True, hide_index=True)
        with b:
            st.subheader("주요 기증처 매출")
            st.bar_chart(donor_df.groupby("기증처", as_index=False)["실매출액"].sum().set_index("기증처"))
            st.dataframe(donor_df, use_container_width=True, hide_index=True)

        st.subheader("분류별 전년 비교")
        st.dataframe(yoy, use_container_width=True, hide_index=True)

with tab4:
    st.subheader("보고서형 엑셀")
    report_bytes = make_report_book(product_df, daily_df)
    st.download_button(
        "확장분석_보고서형_엑셀.xlsx",
        data=report_bytes,
        file_name="확장분석_보고서형_엑셀.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.caption("참고양식 이미지 시트 포함, 요약대시보드/분류전년비교/영수건수비교/결제수단분석/상하위점포 시트 생성")
